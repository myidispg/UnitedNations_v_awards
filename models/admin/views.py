from datetime import datetime

from flask import Blueprint, request, make_response

import admin
from common.database_about import About
from common.database_education import Education
from common.database_language import Language
from common.database_reference import Reference
from common.utils import Utils
from email_alerts import EmailAlerts
from models.users.user import User
from temp import *
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook

__author__ = 'myidispg'

admin_blueprint = Blueprint('admin', __name__)


@admin_blueprint.route('/save_password', methods=['POST'])
def save_password():
    password = request.form.get('admin_password')
    hashed_password = Utils.hash_password(password)
    admin.ADMIN_PASSWORD = str(hashed_password)

    return str(hashed_password)


@admin_blueprint.route('/auth', methods=['POST', 'GET'])
def login_admin():
    if request.method == 'POST':
        username = request.form.get('admin_username')
        password = request.form.get('admin_password')

        if Utils.check_hashed_password(password, admin.ADMIN_PASSWORD) and username == admin.ADMIN_USERNAME:
            return 'Welcome to admin page'
        else:
            return 'Your credentials were wrong.'
    else:
        return "Page upcoming!!!"


@admin_blueprint.route('/send_email_alerts/email_alerts/form_1')
def email_alerts_form_1():
    # if date is greater than reminder date but less than one day added to the reminder date,
    #  the function will proceed.
    if datetime(form_1_reminder_1_year, form_1_reminder_1_month, form_1_reminder_1_day + 1,
                form_1_reminder_1_hour, form_1_reminder_1_minute, form_1_reminder_1_seconds) > \
            datetime.now() > \
            datetime(form_1_reminder_1_year, form_1_reminder_1_month, form_1_reminder_1_day,
                     form_1_reminder_1_hour, form_1_reminder_1_minute, form_1_reminder_1_seconds):
        EmailAlerts.email_alerts_form_1('reminder_1')
    elif datetime(form_1_reminder_2_year, form_1_reminder_2_month, form_1_reminder_2_day + 1,
                  form_1_reminder_2_hour, form_1_reminder_2_minute, form_1_reminder_2_seconds) > \
            datetime.now() > \
            datetime(form_1_reminder_2_year, form_1_reminder_2_month, form_1_reminder_2_day,
                     form_1_reminder_2_hour, form_1_reminder_2_minute, form_1_reminder_2_seconds):
        EmailAlerts.email_alerts_form_1('reminder_2')
    else:
        pass

    return 'emails sent'


@admin_blueprint.route('/send_email_alerts/email_alerts/form_2')
def email_alerts_form_2():
    # if date is greater than reminder date but less than one day added to the reminder date,
    # the function will proceed.
    if datetime(form_2_reminder_1_year, form_2_reminder_1_month, form_2_reminder_1_day + 1,
                form_2_reminder_1_hour, form_2_reminder_1_minute, form_2_reminder_1_seconds) > \
            datetime.now() > \
            datetime(form_2_reminder_1_year, form_2_reminder_1_month, form_2_reminder_1_day,
                     form_2_reminder_1_hour, form_2_reminder_1_minute, form_2_reminder_2_seconds):
        EmailAlerts.email_alerts_form_2('reminder_1')
    elif datetime(form_2_reminder_2_year, form_2_reminder_2_month, form_2_reminder_2_day + 1,
                  form_2_reminder_2_hour, form_2_reminder_2_minute, form_2_reminder_2_seconds) > \
            datetime.now() > \
            datetime(form_2_reminder_2_year, form_2_reminder_2_month, form_2_reminder_2_day,
                     form_2_reminder_2_hour, form_2_reminder_2_minute, form_2_reminder_2_seconds):
        EmailAlerts.email_alerts_form_2('reminder_2')
    else:
        pass

    return 'emails sent'


@admin_blueprint.route('/csv_form_1')
def csv_form_1():
    """
    Generates a excel file with form 1 details of all the users in the database.
    It gets a list of all the users and uses the id of the users to find the necessary details from the
     list of all other fields.
    :return: returns a response object of the excel file which is downloaded.
    """
    user_list = User.get_all_users()
    about_list = About.get_all()
    education_list = Education.get_all()
    reference_list = Reference.get_all()
    language_list = Language.get_all()

    wb = openpyxl.Workbook()
    personal = wb.create_sheet('Personal Info')
    # Create headers
    personal.append(['Email', 'Name', 'Mobile', 'Telephone no', 'Gender', 'DOB', 'Current Address',
                     'Permanent Address', 'Nationality', 'Disability'])
    for user in user_list:
        row_list = [user['email'], user['name'], user['mobile'], user['tel_no'], user['gender'],
                    user['dob'], user['current_address'], user['permanent_address'],
                    user['nationality'], user['disability']]

        personal.append(row_list)

    about = wb.create_sheet('About person')
    # Create headers
    about.append(['User Email', 'About User', 'Why volunteer?', 'Communities Associated with?', 'Motivation'])
    for i in about_list:
        user = User.get_user_object(_id=i['id'])
        row_list = [user.email, i['about'], i['why_volunteer'], i['communities_associated'], i['motivation']]
        about.append(row_list)

    education = wb.create_sheet('Education')
    # Create headers
    education.append(['User Email', 'Course', 'From Date', 'Till Date', 'Institution', 'Board or university'])
    for i in education_list:
        user = User.get_user_object(_id=i['id'])
        row_list = [user.email, i['course'], i['from_date'], i['till_date'], i['institution'], i['board']]
        education.append(row_list)

    reference = wb.create_sheet('References')
    # Create headers
    reference.append(['User Email', 'Reference Name', 'Address', 'Telephone number', 'Reference EmailEmail', 'Occupation',
                      'Relation'])
    for i in reference_list:
        user = User.get_user_object(_id=i['id'])
        row_list = [user.email, i['name'], i['address'], i['tel_no'], i['email'], i['occupation'],
                    i['relation']]
        reference.append(row_list)

    language = wb.create_sheet('Language Info.')
    # Create headers
    language.append(['User Email', 'Language', 'Understand', 'Speaks?', 'Can read or write?'])
    for i in language_list:
        user = User.get_user_object(_id=i['id'])
        row_list = [user.email, i['language'], i['understand'], i['speak'], i['read_write']]
        language.append(row_list)

    # remove the default worksheet from the workbook which is not used.
    wb.remove(wb.worksheets[0])

    response = make_response(save_virtual_workbook(wb))
    response.headers["Content-Disposition"] = "attachment; filename=form_1.xlsx"
    response.headers["Content-type"] = "text/csv"

    return response
